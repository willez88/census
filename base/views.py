import datetime
from tempfile import NamedTemporaryFile

from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.loader import render_to_string
from django.views.generic import TemplateView, View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from weasyprint import CSS, HTML
from weasyprint.text.fonts import FontConfiguration

from user.models import (
    CommunityLeader,
    FamilyGroup,
    Person,
    StreetLeader,
)

from .models import (
    Block,
    Building,
    Department,
    Gender,
    Relationship,
    VoteType,
)


class HomeView(TemplateView):
    """!
    Clase que muestra la página inicial

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/base.html'


class Error403View(TemplateView):
    """!
    Clase que muestra la página de error de permisos

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/error_403.html'


class ExportExcelView(View):
    """!
    Clase que descarga datos relacionados a los usuarios Líder de Comunidad

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-3.0.html'>
        GNU Public License versión 3 (GPLv3)</a>
    """

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        if self.request.user.groups.filter(name='Líder de Comunidad'):
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo excel

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo excel
        """

        workbook = Workbook()
        worksheet1 = workbook.active
        worksheet1.title = 'Hoja 1'
        worksheet1.merge_cells('A1:H1')
        worksheet1.merge_cells('A2:H2')
        worksheet1.merge_cells('A3:H3')
        worksheet1.column_dimensions['A'].width = 20
        worksheet1[
            'A1'
        ] = 'VICEPRESIDENCIA TERRITORIAL PSUV ESTADOS MÉRIDA - TRUJILLO'
        worksheet1['A1'].font = Font(color='FF0000')
        worksheet1['A2'] = 'EQUIPO POLÍTICO ESTADAL PSUV MÉRIDA'
        worksheet1['A2'].font = Font(color='FF0000')
        worksheet1['A3'] = 'COMISIÓN DE ORGANIZACIÓN PSUV MÉRIDA'
        worksheet1['A3'].font = Font(color='FF0000')
        worksheet1['A6'] = 'CENSO RAAS'

        date = datetime.datetime.now()
        worksheet1['A8'] = 'FECHA DE ACTUALIZACIÓN: ' + '%s-%s-%s %s-%s Hrs' %\
            (date.day, date.month, date.year, date.hour, date.minute)

        community_leader = CommunityLeader.objects.get(
            profile=self.request.user.profile
        )
        worksheet1['A10'] = 'Municipio:'
        worksheet1[
            'B10'
        ] = str(community_leader.communal_council.ubch.parish.municipality)

        worksheet1['A11'] = 'Parroquia:'
        worksheet1['B11'] = str(community_leader.communal_council.ubch.parish)

        worksheet1['A12'] = 'Código UBCH:'
        worksheet1['B12'] = community_leader.communal_council.ubch.code

        worksheet1['A13'] = 'Nombre UBCH:'
        worksheet1['B13'] = community_leader.communal_council.ubch.name

        worksheet1['A14'] = 'Código Comunidad:'

        worksheet1['A15'] = 'Nombre Comunidad:'
        worksheet1['B15'] = community_leader.communal_council.name

        worksheet1['A17'] = 'Calles Registradas:'
        worksheet1['B17'] = str(len(StreetLeader.objects.filter(
            community_leader=community_leader))
        )

        # Hojas con los censos
        i = 2
        for street_leader in StreetLeader.objects.filter(
            community_leader=community_leader
        ):
            worksheet = workbook.create_sheet(title='Hoja ' + str(i))
            worksheet.merge_cells('A1:H1')
            worksheet.merge_cells('A2:H2')
            worksheet.merge_cells('A3:H3')
            worksheet.column_dimensions['A'].width = 16
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 16
            worksheet.column_dimensions['D'].width = 25
            worksheet.column_dimensions['E'].width = 20
            worksheet.column_dimensions['F'].width = 20
            worksheet.column_dimensions['G'].width = 20
            worksheet[
                'A1'
            ] = 'VICEPRESIDENCIA TERRITORIAL PSUV ESTADOS MÉRIDA - TRUJILLO'
            worksheet['A1'].font = Font(color='FF0000')
            worksheet['A2'] = 'EQUIPO POLÍTICO ESTADAL PSUV MÉRIDA'
            worksheet['A2'].font = Font(color='FF0000')
            worksheet['A3'] = 'COMISIÓN DE ORGANIZACIÓN PSUV MÉRIDA'
            worksheet['A3'].font = Font(color='FF0000')
            worksheet['A6'] = 'CENSO CALLE RAAS'

            date = datetime.datetime.now()
            worksheet[
                'A8'
            ] = 'FECHA DE ACTUALIZACIÓN: ' + '%s-%s-%s %s-%s Hrs' % \
                (date.day, date.month, date.year, date.hour, date.minute)

            community_leader = CommunityLeader.objects.get(
                profile=self.request.user.profile
            )
            worksheet['A10'] = 'Municipio:'
            worksheet[
                'B10'
            ] = str(community_leader.communal_council.ubch.parish.municipality)

            worksheet['A11'] = 'Parroquia:'
            worksheet[
                'B11'
            ] = str(community_leader.communal_council.ubch.parish)

            worksheet['A12'] = 'Código UBCH:'
            worksheet['B12'] = community_leader.communal_council.ubch.code

            worksheet['A13'] = 'Nombre UBCH:'
            worksheet['B13'] = community_leader.communal_council.ubch.name

            worksheet['A14'] = 'Código Comunidad:'

            worksheet['A15'] = 'Nombre Comunidad:'
            worksheet['B15'] = community_leader.communal_council.name

            worksheet['A16'] = 'Lider de Calle:'
            worksheet['B16'] = str(street_leader.profile)

            worksheet.merge_cells('A19:H19')
            worksheet.merge_cells('A20:H20')
            worksheet['A19'] = 'CENSO REGISTRADO'
            worksheet['A19'].alignment = Alignment(horizontal='center')
            worksheet['A20'] = 'NOTA: TIPO DE VOTOS PERMITIDOS: DURO, BLANDO, \
                OPOSITOR | ESTATUS DE CONTACTADO: SI O NO | JEFE DE FAMILIA: \
                SI O NO | RECIBIÓ CLAP EN LOS ÚLTIMOS TRES MESES: SI O NO'

            worksheet[
                'A21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['A21'].font = Font(color='FFFFFF')
            worksheet['A21'].alignment = Alignment(horizontal='center')

            worksheet[
                'B21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['B21'].font = Font(color='FFFFFF')
            worksheet['B21'].alignment = Alignment(horizontal='center')

            worksheet[
                'C21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['C21'].font = Font(color='FFFFFF')
            worksheet['C21'].alignment = Alignment(horizontal='center')

            worksheet[
                'D21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['D21'].font = Font(color='FFFFFF')
            worksheet['D21'].alignment = Alignment(horizontal='center')

            worksheet[
                'E21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['E21'].font = Font(color='FFFFFF')
            worksheet['E21'].alignment = Alignment(horizontal='center')

            worksheet[
                'F21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['F21'].font = Font(color='FFFFFF')
            worksheet['F21'].alignment = Alignment(horizontal='center')

            worksheet[
                'G21'
            ].fill = PatternFill(start_color='FF0000', fill_type='solid')
            worksheet['G21'].font = Font(color='FFFFFF')
            worksheet['G21'].alignment = Alignment(horizontal='center')

            worksheet['A21'] = 'CÉDULA'
            worksheet['B21'] = 'NOMBRES Y APELLIDOS'
            worksheet['C21'] = 'TELÉFONO'
            worksheet['D21'] = 'Correo'
            worksheet['E21'] = 'TIPO DE VOTO'
            worksheet['F21'] = 'PARENTESCO'
            worksheet['G21'] = 'ES JEFE DE FAMILIA'
            worksheet['H21'] = 'Apartamento'
            c = 22
            for family_group in FamilyGroup.objects.filter(
                street_leader=street_leader
            ):
                # print(Person.objects.filter(family_group=family_group))
                for person in Person.objects.filter(family_group=family_group):
                    column1 = 'A'+str(c)
                    worksheet[column1] = person.id_number

                    column2 = 'B'+str(c)
                    worksheet[
                        column2
                    ] = person.first_name + ' ' + person.last_name

                    column3 = 'C'+str(c)
                    worksheet[column3] = person.phone

                    column4 = 'D'+str(c)
                    worksheet[column4] = person.email

                    column5 = 'E'+str(c)
                    worksheet[column5] = str(person.vote_type)

                    column6 = 'F'+str(c)
                    worksheet[column6] = str(person.relationship)

                    column7 = 'G'+str(c)
                    column8 = 'h'+str(c)
                    if person.family_head:
                        worksheet[column7] = 'SI'
                        worksheet[column8] = str(family_group.department)
                    else:
                        worksheet[column7] = 'No'

                    c = c + 1
                column8 = 'A'+str(c)
                worksheet[column8] = ' '
                c = c + 1

            i = i + 1

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        response = HttpResponse(
            content=stream,
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename="censo.xlsx"'
        # response.write(u'\ufeff'.encode('utf8'))
        return response


class ExportExcelStreetLeaderView(View):
    """!
    Clase que descarga datos relacionados a los usuarios Líder de Calle

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-3.0.html'>
        GNU Public License versión 3 (GPLv3)</a>
    """

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        if self.request.user.groups.filter(name='Líder de Calle'):
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo excel

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo excel
        """

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Hoja 1'
        worksheet.merge_cells('A1:H1')
        worksheet.merge_cells('A2:H2')
        worksheet.merge_cells('A3:H3')
        worksheet.column_dimensions['A'].width = 16
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 16
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20
        worksheet.column_dimensions['G'].width = 20
        worksheet[
            'A1'
        ] = 'VICEPRESIDENCIA TERRITORIAL PSUV ESTADOS MÉRIDA - TRUJILLO'
        worksheet['A1'].font = Font(color='FF0000')
        worksheet['A2'] = 'EQUIPO POLÍTICO ESTADAL PSUV MÉRIDA'
        worksheet['A2'].font = Font(color='FF0000')
        worksheet['A3'] = 'COMISIÓN DE ORGANIZACIÓN PSUV MÉRIDA'
        worksheet['A3'].font = Font(color='FF0000')
        worksheet['A6'] = 'CENSO CALLE RAAS'

        date = datetime.datetime.now()
        worksheet[
            'A8'
        ] = 'FECHA DE ACTUALIZACIÓN: ' + '%s-%s-%s %s-%s Hrs' % \
            (date.day, date.month, date.year, date.hour, date.minute)

        street_leader = StreetLeader.objects.get(
            profile=self.request.user.profile
        )
        community_leader = street_leader.community_leader
        worksheet['A10'] = 'Municipio:'
        worksheet[
            'B10'
        ] = str(community_leader.communal_council.ubch.parish.municipality)

        worksheet['A11'] = 'Parroquia:'
        worksheet[
            'B11'
        ] = str(community_leader.communal_council.ubch.parish)

        worksheet['A12'] = 'Código UBCH:'
        worksheet['B12'] = community_leader.communal_council.ubch.code

        worksheet['A13'] = 'Nombre UBCH:'
        worksheet['B13'] = community_leader.communal_council.ubch.name

        worksheet['A14'] = 'Código Comunidad:'

        worksheet['A15'] = 'Nombre Comunidad:'
        worksheet['B15'] = community_leader.communal_council.name

        worksheet['A16'] = 'Lider de Calle:'
        worksheet['B16'] = str(street_leader.profile)

        worksheet.merge_cells('A19:H19')
        worksheet.merge_cells('A20:H20')
        worksheet['A19'] = 'CENSO REGISTRADO'
        worksheet['A19'].alignment = Alignment(horizontal='center')
        worksheet['A20'] = 'NOTA: TIPO DE VOTOS PERMITIDOS: DURO, BLANDO, \
                OPOSITOR | ESTATUS DE CONTACTADO: SI O NO | JEFE DE FAMILIA: \
                SI O NO | RECIBIÓ CLAP EN LOS ÚLTIMOS TRES MESES: SI O NO'

        worksheet[
            'A21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['A21'].font = Font(color='FFFFFF')
        worksheet['A21'].alignment = Alignment(horizontal='center')

        worksheet[
            'B21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['B21'].font = Font(color='FFFFFF')
        worksheet['B21'].alignment = Alignment(horizontal='center')

        worksheet[
            'C21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['C21'].font = Font(color='FFFFFF')
        worksheet['C21'].alignment = Alignment(horizontal='center')

        worksheet[
            'D21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['D21'].font = Font(color='FFFFFF')
        worksheet['D21'].alignment = Alignment(horizontal='center')

        worksheet[
            'E21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['E21'].font = Font(color='FFFFFF')
        worksheet['E21'].alignment = Alignment(horizontal='center')

        worksheet[
            'F21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['F21'].font = Font(color='FFFFFF')
        worksheet['F21'].alignment = Alignment(horizontal='center')

        worksheet[
            'G21'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['G21'].font = Font(color='FFFFFF')
        worksheet['G21'].alignment = Alignment(horizontal='center')

        worksheet['A21'] = 'CÉDULA'
        worksheet['B21'] = 'NOMBRES Y APELLIDOS'
        worksheet['C21'] = 'TELÉFONO'
        worksheet['D21'] = 'Correo'
        worksheet['E21'] = 'TIPO DE VOTO'
        worksheet['F21'] = 'PARENTESCO'
        worksheet['G21'] = 'ES JEFE DE FAMILIA'
        c = 22
        for family_group in FamilyGroup.objects.filter(
            street_leader=street_leader
        ):
            for person in Person.objects.filter(family_group=family_group):
                column1 = 'A'+str(c)
                worksheet[column1] = person.id_number

                column2 = 'B'+str(c)
                worksheet[
                    column2
                ] = person.first_name + ' ' + person.last_name

                column3 = 'C'+str(c)
                worksheet[column3] = person.phone

                column4 = 'D'+str(c)
                worksheet[column4] = person.email

                column5 = 'E'+str(c)
                worksheet[column5] = str(person.vote_type)

                column6 = 'F'+str(c)
                worksheet[column6] = str(person.relationship)

                column7 = 'G'+str(c)
                if person.family_head:
                    worksheet[column7] = 'SI'
                else:
                    worksheet[column7] = 'No'

                c = c + 1

            column8 = 'A'+str(c)
            worksheet[column8] = ' '
            c = c + 1

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        response = HttpResponse(
            content=stream,
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename="censo.xlsx"'
        return response


class VoteTypeListView(View):
    """!
    Clase que retorna un json con los datos de tipos de voto

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    def get(self, request, *args, **kwargs):
        vote_types = VoteType.objects.all()
        vote_type_list = []
        vote_type_list.append({
            'id': '', 'text': 'Seleccione...'
        })
        for vote_type in vote_types:
            vote_type_list.append({
                'id': vote_type.id, 'text': vote_type.name
            })
        return JsonResponse(
            {'status': 'true', 'list': vote_type_list}, status=200
        )


class RelationshipListView(View):
    """!
    Clase que retorna un json con los datos de parentescos

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    def get(self, request, *args, **kwargs):
        relationships = Relationship.objects.all()
        relationship_list = []
        relationship_list.append({
            'id': '', 'text': 'Seleccione...'
        })
        for relationship in relationships:
            relationship_list.append({
                'id': relationship.id, 'text': relationship.name
            })
        return JsonResponse(
            {'status': 'true', 'list': relationship_list}, status=200
        )


class BuildingListView(View):
    """!
    Clase que retorna un json con los datos de edificios

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    def get(self, request, *args, **kwargs):
        profile = self.request.user.profile
        street_leader = StreetLeader.objects.get(profile=profile)
        buildings = Building.objects.filter(bridge=street_leader.bridge)
        building_list = []
        building_list.append({
            'id': '', 'text': 'Seleccione...'
        })
        for building in buildings:
            building_list.append({
                'id': building.id, 'text': building.name
            })
        return JsonResponse(
            {'status': 'true', 'list': building_list}, status=200
        )


class DepartmentListView(View):
    """!
    Clase que retorna un json con los datos de departamentos

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    def get(self, request, *args, **kwargs):
        departments = Department.objects.all()
        department_list = []
        department_list.append({
            'id': '', 'text': 'Seleccione...'
        })
        for department in departments:
            department_list.append({
                'id': department.id, 'text': department.name
            })
        return JsonResponse(
            {'status': 'true', 'list': department_list}, status=200
        )


class GetDepartmentView(View):
    """!
    Clase que retorna un json con los datos de un departamento

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    def get(self, request, *args, **kwargs):
        """!
        Retorna el json de departamentos filtrados por edificio

        @author William Páez (paez.william8 at gmail.com)
        """

        building = Building.objects.get(pk=self.kwargs['building_id'])
        departments = Department.objects.filter(building=building)
        department_list = []
        department_list.append({
            'id': '', 'text': 'Seleccione...'
        })
        for department in departments:
            department_list.append({
                'id': department.id, 'text': department.name
            })
        return JsonResponse(
            {'status': 'true', 'list': department_list}, status=200
        )


class GenderListView(View):
    """!
    Clase que retorna un json con los datos de géneros

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    def get(self, request, *args, **kwargs):
        genders = Gender.objects.all()
        gender_list = []
        gender_list.append({
            'id': '', 'text': 'Seleccione...'
        })
        for gender in genders:
            gender_list.append({
                'id': gender.id, 'text': gender.name
            })
        return JsonResponse(
            {'status': 'true', 'list': gender_list}, status=200
        )


class VoterTemplateView(TemplateView):
    """!
    Clase que exporta un pdf con votantes mayores o iguales a 15 años

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/voter.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Metodo que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Redirecciona al usuario a la página de error de permisos si no
            es su perfil
        """

        group1 = self.request.user.groups.filter(name='Líder de Comunidad')
        group2 = self.request.user.groups.filter(name='Líder de Calle')
        if group1 or group2:
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        age = int(request.GET.get('age'))
        if CommunityLeader.objects.filter(profile__user=self.request.user):
            community_leader = CommunityLeader.objects.get(profile__user=self.request.user)
            people = Person.objects.filter(
                family_group__street_leader__community_leader=community_leader
            )
        elif StreetLeader.objects.filter(profile__user=self.request.user):
            street_leader = StreetLeader.objects.get(profile__user=self.request.user)
            people = Person.objects.filter(
                family_group__street_leader=street_leader
            )
        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=votantes.pdf'
        font_config = FontConfiguration()
        context = {}
        person_list = []
        for person in people:
            if person.age() >= age:
                person_list.append(person)
        context['people'] = person_list
        html = render_to_string(self.template_name, context)
        HTML(string=html).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
        )
        return response


class DemographicCensusTemplateView(TemplateView):
    """!
    Clase que exporta el censo demográfico

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/demographic_census.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        if self.request.user.groups.filter(name='Líder de Comunidad'):
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=censo-demografico.pdf'
        font_config = FontConfiguration()
        context = {}
        census = []
        for block in Block.objects.all():
            # Total de familias
            families = Person.objects.filter(
                family_head=True,
                family_group__department__building__bridge__block=block
            ).count()
            # Total de personas
            people = Person.objects.filter(
                family_group__department__building__bridge__block=block
            ).count()
            # Total de viviendas
            departments = Person.objects.filter(
                family_head=True,
                family_group__department__building__bridge__block=block
            )
            departments_unique = {}
            for department in departments:
                departments_unique[department.family_group.department] = department
            departments = len(departments_unique)
            MALE = 1
            FEMALE = 2
            # Calcular hembras mayores a 15 años
            females = Person.objects.filter(
                gender__id=FEMALE,
                family_group__department__building__bridge__block=block
            )
            i = 0
            for female in females:
                if female.age() > 15:
                    i = i + 1
            # Total de hembras mayores a 15 años
            females_gt_15 = i
            # Calcular hembras menores a 15 años
            females = Person.objects.filter(
                gender__id=FEMALE,
                family_group__department__building__bridge__block=block
            )
            i = 0
            for female in females:
                if female.age() < 15:
                    i = i + 1
            # Total de hembras menores a 15 años
            females_lt_15 = i

            # Calcular varones mayores a 15 años
            males = Person.objects.filter(
                gender__id=MALE,
                family_group__department__building__bridge__block=block
            )
            i = 0
            for male in males:
                if male.age() > 15:
                    i = i + 1
            # Total de varones mayores a 15 años
            males_gt_15 = i
            # Calcular varones menores a 15 años
            males = Person.objects.filter(
                gender__id=MALE,
                family_group__department__building__bridge__block=block
            )
            i = 0
            for male in males:
                if male.age() < 15:
                    i = i + 1
            # Total de varones menores a 15 años
            males_lt_15 = i

            census.append({
                'block': block.name,
                'families': families,
                'people': people,
                'departments': departments,
                'females_gt_15': females_gt_15,
                'females_lt_15': females_lt_15,
                'males_gt_15': males_gt_15,
                'males_lt_15': males_lt_15,
            })
        context['census'] = census
        html = render_to_string(self.template_name, context)
        HTML(string=html).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
        )
        return response


class VacationPlanTemplateView(TemplateView):
    """!
    Clase que exporta niños entre 7 y 12 años de edad

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/vacation_plan.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        if self.request.user.groups.filter(name='Líder de Comunidad'):
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=plan-vacacional.pdf'
        font_config = FontConfiguration()
        context = {}
        childrens = []
        for person in Person.objects.all():
            if person.age() >= 7 and person.age() <= 12:
                family_group = FamilyGroup.objects.get(person=person)
                for person2 in family_group.person_set.all():
                    if person2.family_head:
                        childrens.append({
                            'family_head': person2,
                            'children': person
                        })
        context['people'] = childrens
        html = render_to_string(self.template_name, context)
        HTML(string=html).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
        )
        return response


class FilterTemplateView(TemplateView):
    """!
    Clase que muestra página de filtros

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/filter.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Metodo que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Redirecciona al usuario a la página de error de permisos si no
            es su perfil
        """

        group1 = self.request.user.groups.filter(name='Líder de Comunidad')
        group2 = self.request.user.groups.filter(name='Líder de Calle')
        if group1 or group2:
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')


class FilterAgeTemplateView(TemplateView):
    """!
    Clase que exporta pdf de niños entre 2 edades

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/filter_age.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Metodo que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Redirecciona al usuario a la página de error de permisos si no
            es su perfil
        """

        group1 = self.request.user.groups.filter(name='Líder de Comunidad')
        group2 = self.request.user.groups.filter(name='Líder de Calle')
        if group1 or group2:
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        age1 = int(request.GET.get('age1'))
        age2 = int(request.GET.get('age2'))
        if CommunityLeader.objects.filter(profile__user=self.request.user):
            community_leader = CommunityLeader.objects.get(profile__user=self.request.user)
            communal_council = community_leader.communal_council
            people = Person.objects.filter(
                family_group__street_leader__community_leader__communal_council=communal_council
            )
        elif StreetLeader.objects.filter(profile__user=self.request.user):
            street_leader = StreetLeader.objects.get(profile__user=self.request.user)
            people = Person.objects.filter(
                family_group__street_leader=street_leader
            )
        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=edades.pdf'
        font_config = FontConfiguration()
        context = {}
        childrens = []
        for person in people:
            if person.age() >= age1 and person.age() <= age2:
                family_group = FamilyGroup.objects.get(person=person)
                for person2 in family_group.person_set.all():
                    if person2.family_head:
                        childrens.append({
                            'family_head': person2,
                            'children': person
                        })
        context['people'] = childrens
        html = render_to_string(self.template_name, context)
        HTML(string=html).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
        )
        return response


class SociodemographicTemplateView(TemplateView):
    """!
    Clase que exporta el censo sociodemográfico

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/sociodemographic.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        if self.request.user.groups.filter(name='Líder de Comunidad'):
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=sociodemografico.pdf'
        font_config = FontConfiguration()
        context = {}
        census = []
        community_leader = CommunityLeader.objects.get(profile__user=self.request.user)
        communal_council = community_leader.communal_council
        # Total de familias
        families = Person.objects.filter(
            family_head=True,
            family_group__department__building__bridge__block__communal_council=communal_council
        ).count()

        # Total de viviendas
        departments = Person.objects.filter(
            family_head=True,
            family_group__department__building__bridge__block__communal_council=communal_council
        )
        departments_unique = {}
        for department in departments:
            departments_unique[department.family_group.department] = department
        departments = len(departments_unique)

        MALE = 1
        # Personas masculinas
        males = Person.objects.filter(
            gender__id=MALE,
            family_group__department__building__bridge__block__communal_council=communal_council
        )
        i = 0
        for male in males:
            if male.age() >= 0 and male.age() <= 12:
                i = i + 1
        # Niños
        male_children = i
        i = 0
        for male in males:
            if male.age() >= 13 and male.age() <= 17:
                i = i + 1
        # Adolescentes barones
        male_teen = i
        i = 0
        for male in males:
            if male.age() >= 18:
                i = i + 1
        # Hombres mayores o iguales a 18 años
        male_gte_18 = i
        i = 0
        for male in males:
            if male.age() >= 60:
                i = i + 1
        # Adulto mayor hombres
        male_elderly = i

        FEMALE = 2
        # Personas femeninas
        females = Person.objects.filter(
            gender__id=FEMALE,
            family_group__department__building__bridge__block__communal_council=communal_council
        )
        for female in females:
            if female.age() >= 0 and female.age() <= 12:
                i = i + 1
        # Niñas
        female_children = i
        i = 0
        for female in females:
            if female.age() >= 13 and female.age() <= 17:
                i = i + 1
        # Adolescentes hembras
        female_teen = i
        i = 0
        for female in females:
            if female.age() >= 18:
                i = i + 1
        # Mujeres mayores o iguales a 18 años
        female_gte_18 = i
        i = 0
        for female in females:
            if female.age() >= 55:
                i = i + 1
        # Adulto mayor Hembras
        female_elderly = i

        census.append({
                'departments': departments,
                'families': families,
                'male_children': male_children,
                'male_teen': male_teen,
                'male_gte_18': male_gte_18,
                'male_elderly': male_elderly,
                'female_children': female_children,
                'female_teen': female_teen,
                'female_gte_18': female_gte_18,
                'female_elderly': female_elderly,
            })
        context['census'] = census
        html = render_to_string(self.template_name, context)
        HTML(string=html).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
        )
        return response


class ResidenceProofTemplateView(TemplateView):
    """!
    Clase que exporta la constancia de residencia

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/residence_proof.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        group1 = self.request.user.groups.filter(name='Líder de Comunidad')
        group2 = self.request.user.groups.filter(name='Líder de Calle')
        group3 = self.request.user.groups.filter(name='Grupo Familiar')
        if group1 or group2 or group3:
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        id_number = kwargs['id_number']
        person = get_object_or_404(Person, id_number=id_number)
        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=carta-residencia.pdf'
        font_config = FontConfiguration()
        context = {}
        context['person'] = person
        context['logo_url'] = settings.BASE_DIR / 'static/img/logo-rdsr.jpg'
        context['imagen1'] = settings.BASE_DIR / 'static/img/imagen1.png'
        context['mara'] = settings.BASE_DIR / 'static/img/mara.png'
        context['jairo'] = settings.BASE_DIR / 'static/img/jairo.png'
        html = render_to_string(self.template_name, context)
        HTML(
            string=html, base_url=request.build_absolute_uri()
        ).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
            presentational_hints=True
        )
        return response


class LowResourcesTemplateView(TemplateView):
    """!
    Clase que exporta la carta de bajos recursos

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-2.0.html'>
        GNU Public License versión 2 (GPLv2)</a>
    """

    template_name = 'base/low_resources.html'

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        group1 = self.request.user.groups.filter(name='Líder de Comunidad')
        group2 = self.request.user.groups.filter(name='Líder de Calle')
        group3 = self.request.user.groups.filter(name='Grupo Familiar')
        if group1 or group2 or group3:
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')

    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo pdf

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo pdf
        """

        id_number = kwargs['id_number']
        person = get_object_or_404(Person, id_number=id_number)
        response = HttpResponse(content_type='application/pdf')
        response[
            'Content-Disposition'
        ] = 'inline; filename=carta-residencia.pdf'
        font_config = FontConfiguration()
        context = {}
        context['person'] = person
        context['logo_url'] = settings.BASE_DIR / 'static/img/logo-rdsr.jpg'
        context['imagen1'] = settings.BASE_DIR / 'static/img/imagen1.png'
        context['mara'] = settings.BASE_DIR / 'static/img/mara.png'
        context['jairo'] = settings.BASE_DIR / 'static/img/jairo.png'
        html = render_to_string(self.template_name, context)
        HTML(
            string=html, base_url=request.build_absolute_uri()
        ).write_pdf(
            response,
            font_config=font_config,
            stylesheets=[
                CSS(settings.BASE_DIR / 'static/css/bootstrap.min.css')
            ],
            presentational_hints=True
        )
        return response


class ExportExcelOlderAdultView(View):
    """!
    Clase que descarga adultos mayores relacionados a los usuarios Líder de Comunidad

    @author William Páez (paez.william8 at gmail.com)
    @copyright <a href='http://www.gnu.org/licenses/gpl-3.0.html'>
        GNU Public License versión 3 (GPLv3)</a>
    """

    def dispatch(self, request, *args, **kwargs):
        """!
        Función que valida si el usuario del sistema tiene permisos para entrar
        a esta vista

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene los datos de la
            petición
        @param *args <b>{tuple}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return super <b>{object}</b> Entra a la vista correspondiente
            sino redirecciona hacia la vista de error de permisos
        """

        if self.request.user.groups.filter(name='Líder de Comunidad'):
            return super().dispatch(request, *args, **kwargs)
        return redirect('base:error_403')
    
    def get(self, request, *args, **kwargs):
        """!
        Función que descarga un archivo excel

        @author William Páez (paez.william8 at gmail.com)
        @param self <b>{object}</b> Objeto que instancia la clase
        @param request <b>{object}</b> Objeto que contiene la petición
        @param *args <b>{tupla}</b> Tupla de valores, inicialmente vacia
        @param **kwargs <b>{dict}</b> Diccionario de datos, inicialmente vacio
        @return Retorna datos en un archivo excel
        """

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Hoja 1'
        community_leader = CommunityLeader.objects.get(
            profile=self.request.user.profile
        )
        worksheet.column_dimensions['A'].width = 30
        worksheet.row_dimensions[1].height = 60
        worksheet[
                'A1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['A1'].font = Font(color='FFFFFF', bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['B'].width = 30
        worksheet[
            'B1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['B1'].font = Font(color='FFFFFF', bold=True)
        worksheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['C'].width = 30
        worksheet[
            'C1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['C1'].font = Font(color='FFFFFF', bold=True)
        worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['D'].width = 30
        worksheet[
            'D1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['D1'].font = Font(color='FFFFFF', bold=True)
        worksheet['D1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['E'].width = 50
        worksheet[
            'E1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['E1'].font = Font(color='FFFFFF', bold=True)
        worksheet['E1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['F'].width = 50
        worksheet[
            'F1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['F1'].font = Font(color='FFFFFF', bold=True)
        worksheet['F1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['G'].width = 50
        worksheet[
            'G1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['G1'].font = Font(color='FFFFFF', bold=True)
        worksheet['G1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['H'].width = 30
        worksheet[
            'H1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['H1'].font = Font(color='FFFFFF', bold=True)
        worksheet['H1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['I'].width = 50
        worksheet[
            'I1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['I1'].font = Font(color='FFFFFF', bold=True)
        worksheet['I1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['J'].width = 50
        worksheet[
            'J1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['J1'].font = Font(color='FFFFFF', bold=True)
        worksheet['J1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['K'].width = 30
        worksheet[
            'K1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['K1'].font = Font(color='FFFFFF', bold=True)
        worksheet['K1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['L'].width = 30
        worksheet[
            'L1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['L1'].font = Font(color='FFFFFF', bold=True)
        worksheet['L1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['M'].width = 30
        worksheet[
            'M1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['M1'].font = Font(color='FFFFFF', bold=True)
        worksheet['M1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['N'].width = 30
        worksheet[
            'N1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['N1'].font = Font(color='FFFFFF', bold=True)
        worksheet['N1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['O'].width = 30
        worksheet[
            'O1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['O1'].font = Font(color='FFFFFF', bold=True)
        worksheet['O1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['P'].width = 30
        worksheet[
            'P1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['P1'].font = Font(color='FFFFFF', bold=True)
        worksheet['P1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['Q'].width = 30
        worksheet[
            'Q1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['Q1'].font = Font(color='FFFFFF', bold=True)
        worksheet['Q1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['R'].width = 30
        worksheet[
            'R1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['R1'].font = Font(color='FFFFFF', bold=True)
        worksheet['R1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['S'].width = 30
        worksheet[
            'S1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['S1'].font = Font(color='FFFFFF', bold=True)
        worksheet['S1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.column_dimensions['T'].width = 30
        worksheet[
            'T1'
        ].fill = PatternFill(start_color='FF0000', fill_type='solid')
        worksheet['T1'].font = Font(color='FFFFFF', bold=True)
        worksheet['T1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet['A1'] = 'Municipio'
        worksheet['B1'] = 'Parroquia'
        worksheet['C1'] = 'Nombre del CLAP'
        worksheet['D1'] = 'Comunidades Asociadas al CLAP'
        worksheet['E1'] = 'Jefe de Calle'
        worksheet['F1'] = 'Cédula de Identidad del Jefe de Familia (Solo números)'
        worksheet['G1'] = 'Nombres y Apellidos del Jefe de Familia'
        worksheet['H1'] = 'Teléfono'
        worksheet['I1'] = 'Dirección Exacta (Calle/Edif/Avenida,numero)'
        worksheet['J1'] = 'Cantidad de Carga Familiar (incluyendo del Jefe de Familia)'
        worksheet['K1'] = 'Cédula Carga Familiar 1'
        worksheet['L1'] = 'Cédula Carga Familiar 2'
        worksheet['M1'] = 'Cédula Carga Familiar 3'
        worksheet['N1'] = 'Cédula Carga Familiar 4'
        worksheet['O1'] = 'Cédula Carga Familiar 5'
        worksheet['P1'] = 'Cédula Carga Familiar 6'
        worksheet['Q1'] = 'Cédula Carga Familiar 7'
        worksheet['R1'] = 'Cédula Carga Familiar 8'
        worksheet['S1'] = 'Cédula Carga Familiar 9'
        worksheet['T1'] = 'Cédula Carga Familiar 10'
        columns = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
        c = 2
        for street_leader in StreetLeader.objects.filter(
            community_leader=community_leader
        ):
            for family_group in FamilyGroup.objects.filter(
                street_leader=street_leader
            ):
                family_head = family_group.person_set.filter(family_head=True)
                for person in Person.objects.filter(family_group=family_group):
                    if person.age() >= 50:
                        column1 = 'A'+str(c)
                        worksheet[column1] = str(community_leader.communal_council.ubch.parish.municipality)

                        column2 = 'B'+str(c)
                        worksheet[column2] = str(community_leader.communal_council.ubch.parish)
                        
                        column3 = 'C'+str(c)
                        worksheet[column3] = 'Domingo Salazar Rojas'

                        column5 = 'E'+str(c)
                        worksheet[column5] = str(street_leader.profile)

                        column6 = 'F'+str(c)
                        worksheet[column6] = str(family_head[0].id_number)

                        column7 = 'G'+str(c)
                        worksheet[column7] = family_head[0].first_name + ' ' + family_head[0].last_name

                        column8 = 'H'+str(c)
                        worksheet[column8] = family_head[0].phone

                        column9 = 'I'+str(c)
                        worksheet[column9] = str(family_head[0].family_group.department)

                        column10 = 'J'+str(c)
                        worksheet[column10] = family_group.person_set.count()

                        i = 0
                        for person in family_group.person_set.all():
                            worksheet[columns[i]+str(c)] = person.id_number
                            i = i + 1

                        c = c + 1
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        response = HttpResponse(
            content=stream,
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename="censo_estado_mayor.xlsx"'
        return response
